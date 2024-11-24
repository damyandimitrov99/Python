from ftplib import FTP
from datetime import datetime, timezone, timedelta
import pytz
import yfinance as yf
import numpy as np
import pandas as pd
from scipy import signal
import mysql.connector
from bs4 import BeautifulSoup
import asyncio
import aiohttp
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font
import re



# FileZilla server details
host = 'ftp.nasdaqtrader.com'
username = 'anonymous'
password = 'anonymous@example.com'

# Create an FTP object and connect to the server
with FTP(host) as ftp:
    ftp.login(username, password)

    # Change the working directory to 'Symboldirectory'
    ftp.cwd('Symboldirectory')

    # Function to store lines in a list of lists (split by '|')
    file_contents_nasdaq = []
    file_contents_other = []

    def store_line_nasdaq(line):
        file_contents_nasdaq.append(line.split('|'))

    def store_line_other(line):
        file_contents_other.append(line.split('|'))

    # Read the contents of 'nasdaqlisted.txt' and 'otherlisted.txt'
    ftp.retrlines('RETR nasdaqlisted.txt', store_line_nasdaq)
    ftp.retrlines('RETR otherlisted.txt', store_line_other)

# Create DataFrames from the lists of lists with column names taken from the first row
df_nasdaq = pd.DataFrame(file_contents_nasdaq[1:], columns=file_contents_nasdaq[0])
df_other = pd.DataFrame(file_contents_other[1:], columns=file_contents_other[0])

# Rename the 'ACT Symbol' column to 'Symbol'
df_other = df_other.rename(columns={'ACT Symbol': 'Symbol'})

# Remove the last row from df_nasdaq and df_other
df_nasdaq = df_nasdaq.iloc[:-1]
df_other = df_other.iloc[:-1]

# Delete rows where 'ETF' column is 'Y' or where 'Test Issue' column is 'Y'
df_nasdaq = df_nasdaq[(df_nasdaq['ETF'] != 'Y') & (df_nasdaq['Test Issue'] != 'Y')]
df_other = df_other[(df_other['ETF'] != 'Y') & (df_other['Test Issue'] != 'Y')]

# Filter out rows where 'Symbol' contains '.' or '$'
df_other = df_other[~df_other['Symbol'].str.contains('[.$]')]

# Replace values in the 'Market Category' and 'Financial Status' columns
market_category_mapping = {
    'Q': 'NASDAQ Global Select MarketSM',
    'G': 'NASDAQ Global MarketSM',
    'S': 'NASDAQ Capital Market'
}

financial_status_mapping = {
    'D': 'Deficient: Issuer Failed to Meet NASDAQ Continued Listing Requirements',
    'E': 'Delinquent: Issuer Missed Regulatory Filing Deadline',
    'Q': 'Bankrupt: Issuer Has Filed for Bankruptcy',
    'N': 'Normal (Default): Issuer Is NOT Deficient, Delinquent, or Bankrupt',
    'G': 'Deficient and Bankrupt',
    'H': 'Deficient and Delinquent',
    'J': 'Delinquent and Bankrupt',
    'K': 'Deficient, Delinquent, and Bankrupt'
}

df_nasdaq['Market Category'] = df_nasdaq['Market Category'].map(market_category_mapping)
df_nasdaq['Financial Status'] = df_nasdaq['Financial Status'].map(financial_status_mapping)
df_nasdaq = df_nasdaq.drop(columns=['Test Issue', 'ETF', 'Round Lot Size', 'NextShares'])

# Replace values in the 'Exchange' column
exchange_mapping = {
    'A': 'NYSE MKT',
    'N': 'New York Stock Exchange (NYSE)',
    'P': 'NYSE ARCA',
    'Z': 'BATS Global Markets (BATS)',
    'V': "Investors' Exchange, LLC (IEXG)"
}

df_other['Exchange'] = df_other['Exchange'].map(exchange_mapping)
df_other = df_other.drop(columns=['Test Issue', 'ETF', 'Round Lot Size'])

# Create an Excel file with the data from df_nasdaq and df_other
with pd.ExcelWriter('nasdaq_and_other_data.xlsx') as writer:
    df_nasdaq.to_excel(writer, sheet_name='nasdaqlisted', index=False)
    df_other.to_excel(writer, sheet_name='otherlisted', index=False)


# Asynchronous function to get sector
async def get_sector_async(session, symbol):
    try:
        sector_df = pd.DataFrame(columns=['Symbol', 'Sector'], index=None)
        industry_df = pd.DataFrame(columns=['Symbol', 'Industry'], index=None)
        url = f'https://finance.yahoo.com/quote/{symbol}/profile?p={symbol}'
        headers = {'User-agent': 'Mozilla/5.0'}
        async with session.get(url, headers=headers) as response:
            soup = BeautifulSoup(await response.text(), features='lxml')
            sector_tag = soup.find('span', string='Sector(s)').find_next('span')
            industry_tag = soup.find('span', string='Industry').find_next('span')
            sector = sector_tag.text if sector_tag else 'N/A'
            industry = industry_tag.text if industry_tag else 'N/A'
            sector_df = pd.concat([sector_df, pd.DataFrame({'Symbol': [symbol], 'Sector': [sector]})], ignore_index=True)
            industry_df = pd.concat([industry_df, pd.DataFrame({'Symbol': [symbol], 'Industry': [industry]})], ignore_index=True)
            print("Symbol:", symbol)
            print("Sector:", sector,)
            print("Industry:", industry, '\n')
            return sector_df, industry_df
    except AttributeError:
        return False
        print()
        print("Symbol:", symbol)
        print("Sector: not found\n")


# Asynchronous function to get market cap
async def get_market_cap_async(session, symbol):
    try:
        market_cap_df = pd.DataFrame(columns=['Symbol', 'Market Cap'], index=None)
        # print(market_cap_df)
        market_cap_key = 'Market Cap (intraday)'
        url = f'https://finance.yahoo.com/quote/{symbol}/key-statistics?p={symbol}'
        headers = {'User-agent': 'Mozilla/5.0'}
        async with session.get(url, headers=headers) as response:
            soup = BeautifulSoup(await response.text(), features='lxml')
            tables = soup.findAll('table')

            for table in tables:
                for tr in table.find_all('tr'):
                    row = [td.text.strip() for td in tr.find_all('td')]
                    if row and row[0] == market_cap_key:
                        market_cap = row[1]
                        if market_cap == 'N/A':
                            return False
                        if 'M' not in market_cap:
                            print("Symbol:", symbol)
                            print("Market Cap:", market_cap, '\n')
                            market_cap_df = pd.concat([market_cap_df, pd.DataFrame({'Symbol': [symbol], 'Market Cap': [market_cap]})], ignore_index=True)
                            # print(market_cap_df)
                            return market_cap_df
                        if 'M' in market_cap:
                            number_part = float(market_cap[:-1])
                            if number_part < 500:
                                return False
                            else:
                                print("Symbol:", symbol)
                                print("Market Cap:", market_cap, '\n')
                                market_cap_df = pd.concat([market_cap_df, pd.DataFrame({'Symbol': [symbol], 'Market Cap': [market_cap]})], ignore_index=True)
                                # print(market_cap)
                                return market_cap_df

    except Exception as e:
        return False



# Asynchronous function to get the last close price
async def extrema(symbol):
    try:
        current_local_time = datetime.now()
        utc_timezone = pytz.timezone('UTC')
        current_date = current_local_time.astimezone(utc_timezone)
        start_date = current_date - timedelta(days=370)

        timeframe = '1D'
        data = yf.download(symbol, start=start_date, interval=timeframe, progress=False)
        
        end_date = data.index[-2]
        start_date = end_date - timedelta(days=365)

        data = data[(start_date <= data.index) & (data.index <= end_date)]

        c = data['Close']
        o = data['Open']

        h = data['High']
        l = data['Low']



        #EXTREMA 1
        highs_peaks = signal.argrelextrema(h.values, np.greater)[0]
        lows_valleys = signal.argrelextrema(l.values, np.less)[0]

        # High and Low Merge peaks and valleys data points using pandas extrema
        df_highs_peaks = pd.DataFrame({'Date': h.index[highs_peaks], 'Extremum': h.values[highs_peaks]})
        df_lows_valleys = pd.DataFrame({'Date': l.index[lows_valleys], 'Extremum': l.values[lows_valleys]})

        ex1_df_highs_peaks_and_lows_valleys = pd.concat([df_highs_peaks, df_lows_valleys], axis=0, ignore_index=True, sort=True)
        ex1_df_highs_peaks_and_lows_valleys = ex1_df_highs_peaks_and_lows_valleys.sort_values(by=['Date'])

        # Apply the ZigZag filter
        ex1_extrema = []
        ex1_extrema.append((ex1_df_highs_peaks_and_lows_valleys['Date'].iloc[0], ex1_df_highs_peaks_and_lows_valleys['Extremum'].iloc[0]))

        for i in range(len(ex1_df_highs_peaks_and_lows_valleys)):
            if (((ex1_df_highs_peaks_and_lows_valleys.iloc[i][['Date']] == df_highs_peaks[['Date']]).all(axis=1).any()) and ((ex1_extrema[-1][0] == df_highs_peaks[['Date']]).all(axis=1).any())):
                continue

            if (((ex1_df_highs_peaks_and_lows_valleys.iloc[i][['Date']] == df_lows_valleys[['Date']]).all(axis=1).any()) and ((ex1_extrema[-1][0] == df_lows_valleys[['Date']]).all(axis=1).any())):
                continue

            else:
                ex1_extrema.append((ex1_df_highs_peaks_and_lows_valleys['Date'].iloc[i], ex1_df_highs_peaks_and_lows_valleys['Extremum'].iloc[i]))


        # Convert ex1_extrema to DataFrame
        ex1_df_extrema = pd.DataFrame(ex1_extrema, columns=['Date', 'Extremum'])
        ex1_df_extrema = ex1_df_extrema.sort_values(by='Date')



        #EXTREMA 2
        # Convert the 'date' column to a list of datetime objects
        df_list_peaks_dates = df_highs_peaks['Date'].tolist()
        df_list_valleys_dates = df_lows_valleys['Date'].tolist()

        # Create a new DataFrame to store the new data
        ex1_df_highs_peaks = []
        ex1_df_lows_valleys = []

        for date, extremum in ex1_extrema:
            if date in df_list_peaks_dates:
                ex1_df_highs_peaks.append({'Date': date, 'Extremum': extremum})


        for date, extremum in ex1_extrema:
            if date in df_list_valleys_dates:
                ex1_df_lows_valleys.append({'Date': date, 'Extremum': extremum})


        # Convert the ex2_df_highs_peaks list to a DataFrame
        ex1_df_highs_peaks = pd.DataFrame(ex1_df_highs_peaks)
        ex1_df_lows_valleys = pd.DataFrame(ex1_df_lows_valleys)

        # Create a new DataFrame to store the new data
        ex2_df_highs_peaks = []
        ex2_df_lows_valleys = []

        # Iterate through ex1_extrema and add new data to the ex2_df_highs_peaks list
        for date, extremum in ex1_extrema:
            if date in df_list_peaks_dates:
                ex2_df_highs_peaks.append({'Date': date, 'Extremum': extremum})


        for date, extremum in ex1_extrema:
            if date in df_list_valleys_dates:
                ex2_df_lows_valleys.append({'Date': date, 'Extremum': extremum})


        # Convert the ex2_df_highs_peaks list to a DataFrame
        ex2_df_highs_peaks = pd.DataFrame(ex2_df_highs_peaks)
        ex2_df_lows_valleys = pd.DataFrame(ex2_df_lows_valleys)

        # Find local extrema using argrelextrema
        ex2_extrema_highs_peaks = signal.argrelextrema(ex2_df_highs_peaks['Extremum'].values, np.greater, order=1)
        ex2_extrema_lows_valleys = signal.argrelextrema(ex2_df_lows_valleys['Extremum'].values, np.less, order=1)

        # Retrieve dates and values
        ex2_highs_peaks_dates = ex2_df_highs_peaks.loc[ex2_extrema_highs_peaks[0], 'Date']
        ex2_highs_peaks_values = ex2_df_highs_peaks.loc[ex2_extrema_highs_peaks[0], 'Extremum']

        ex2_lows_valleys_dates = ex2_df_lows_valleys.loc[ex2_extrema_lows_valleys[0], 'Date']
        ex2_lows_valleys_values = ex2_df_lows_valleys.loc[ex2_extrema_lows_valleys[0], 'Extremum']

        # Create a DataFrame with extrema information
        ex2_df_extrema_highs_peaks = pd.DataFrame({'Date': ex2_highs_peaks_dates, 'Extremum': ex2_highs_peaks_values})
        ex2_df_extrema_lows_valleys = pd.DataFrame({'Date': ex2_lows_valleys_dates, 'Extremum': ex2_lows_valleys_values})   

        # Concatenate the two dataframes
        ex2_df_highs_peaks_and_lows_valleys = pd.concat([ex2_df_extrema_highs_peaks, ex2_df_extrema_lows_valleys], ignore_index=True)
        ex2_df_highs_peaks_and_lows_valleys = pd.DataFrame(ex2_df_highs_peaks_and_lows_valleys, columns=['Date', 'Extremum'])
        ex2_df_highs_peaks_and_lows_valleys = ex2_df_highs_peaks_and_lows_valleys.sort_values(by='Date')

        ex2_extrema = []
        ex2_df_extrema = []

        #Filter for extrema
        ex2_extrema.append((ex2_df_highs_peaks_and_lows_valleys['Date'].iloc[0], ex2_df_highs_peaks_and_lows_valleys['Extremum'].iloc[0]))

        for i in range(len(ex2_df_highs_peaks_and_lows_valleys)):
            if (((ex2_df_highs_peaks_and_lows_valleys.iloc[i][['Date']] == df_highs_peaks[['Date']]).all(axis=1).any()) and ((ex2_extrema[-1][0] == df_highs_peaks[['Date']]).all(axis=1).any())):
                continue

            if (((ex2_df_highs_peaks_and_lows_valleys.iloc[i][['Date']] == df_lows_valleys[['Date']]).all(axis=1).any()) and ((ex2_extrema[-1][0] == df_lows_valleys[['Date']]).all(axis=1).any())):
                continue

            else:
                ex2_extrema.append((ex2_df_highs_peaks_and_lows_valleys['Date'].iloc[i], ex2_df_highs_peaks_and_lows_valleys['Extremum'].iloc[i]))


        # Convert ex2_df_extrema to DataFrame
        ex2_df_extrema = pd.DataFrame(ex2_extrema, columns=['Date', 'Extremum'])
        ex2_df_extrema = ex2_df_extrema.sort_values(by='Date')

        return ex1_df_extrema, ex2_df_extrema, data
        
    except Exception as e:
        return False



#Check for Database
async def mysql_databse_connection(symbol):
    # Define your MySQL connection parameters
    host = 'localhost'
    user = 'root'
    password = 'Damyan110799'
    database = 'tangra_financial_analyst_stocks'

    conn = mysql.connector.connect(host=host, user=user, password=password)
    cursor = conn.cursor()

    #Check If the DataBase Exists
    cursor.execute('''SHOW DATABASES;''')
    databases = cursor.fetchall()
    tangra_financial_analyst_scalper_exists = any(database in db for db in databases)

    #Check for Existing Database
    if tangra_financial_analyst_scalper_exists:
        conn = mysql.connector.connect(
            host='localhost',
            user='root',
            password='Damyan110799',
            database=database
        )
        cursor = conn.cursor()

    #Check for Not Existing Database
    if not tangra_financial_analyst_scalper_exists:
        cursor.execute(f'''CREATE DATABASE {database};''')
        conn = mysql.connector.connect(
            host='localhost',
            user='root',
            password='Damyan110799',
            database=database
        )
        cursor = conn.cursor()

    #Check If the Table Exists
    cursor.execute(f'''SHOW TABLES LIKE '{symbol}';''')
    table_exists = cursor.fetchall()

    return conn, cursor, table_exists



#Not Existing Table
async def mysql_not_existing_table(symbol, conn, cursor, data, ex1_df_extrema, ex2_df_extrema):
    cursor.execute(f'''
        CREATE TABLE IF NOT EXISTS `{symbol}`(
            Date_Info DATETIME PRIMARY KEY,
            Open FLOAT(10,5),
            High FLOAT(10,5),
            Low FLOAT(10,5),
            Close FLOAT(10,5),
            Extrema_1 FLOAT(10,5),
            Extrema_2 FLOAT(10,5)
            )
    ''')

    for index, row in data.iterrows():
        cursor.execute(f'''
            INSERT INTO `{symbol}`
            (Date_Info, Open, High, Low, Close)
            VALUES (%s, %s, %s, %s, %s)
        ''', (index, row['Open'], row['High'], row['Low'], row['Close']))

    for _, row in ex1_df_extrema.iterrows():
        cursor.execute(f'''
            UPDATE `{symbol}`
            SET Extrema_1 = %s
            WHERE Date_Info = %s;
        ''', (row['Extremum'], row['Date']))

    if len(ex2_df_extrema) > 0:
        for _, row in ex2_df_extrema.iterrows():
            cursor.execute(f'''
                UPDATE `{symbol}`
                SET Extrema_2 = %s
                WHERE Date_Info = %s;
            ''', (row['Extremum'], row['Date']))

    conn.commit()



#Existing Table
async def mysql_existing_table_checking_new_data(symbol, cursor):
    # Execute the SQL query to get the last Date_Info
    cursor.execute(f'''SELECT MAX(Date_Info) FROM `{symbol}`''')
    last_date_info = cursor.fetchone()[0]

    return last_date_info



#Checking For New Data
async def checking_new_data(symbol, last_date_info):
    timeframe = '1D'
    start_date = last_date_info

    data = yf.download(symbol, start=start_date, interval=timeframe, progress=False)

    if len(data)>2:
        return True
    else:
        return False



async def mysql_last_data(symbol, cursor):
    cursor.execute(f'''SELECT Extrema_1 FROM `{symbol}` WHERE Extrema_1 IS NOT NULL ORDER BY Date_Info DESC LIMIT 1''')
    last_extrema_1_value = cursor.fetchone()[0]

    cursor.execute(f'''SELECT Date_Info FROM `{symbol}` WHERE Extrema_1 IS NOT NULL ORDER BY Date_Info DESC LIMIT 1''')
    last_extrema_1_date = cursor.fetchone()[0]

    cursor.execute(f'''SELECT Extrema_2 FROM `{symbol}` WHERE Extrema_2 IS NOT NULL ORDER BY Date_Info DESC LIMIT 1''')
    last_extrema_2_value = cursor.fetchone()[0]

    cursor.execute(f'''SELECT Date_Info FROM `{symbol}` WHERE Extrema_2 IS NOT NULL ORDER BY Date_Info DESC LIMIT 1''')
    last_extrema_2_date = cursor.fetchone()[0]

    return last_extrema_1_value, last_extrema_1_date, last_extrema_2_value, last_extrema_2_date



async def new_extrema(symbol):
    try:
        timeframe = '1D'
        start_date = last_extrema_2_date

        data = yf.download(symbol, start=start_date, interval=timeframe, progress=False)

        #Final Date
        end_date = data.index[-2]
        data = data[(start_date <= data.index) & (data.index <= end_date)]

        c = data['Close']
        o = data['Open']

        h = data['High']
        l = data['Low']



        #EXTREMA 1
        highs_peaks = signal.argrelextrema(h.values, np.greater)[0]
        lows_valleys = signal.argrelextrema(l.values, np.less)[0]

        # High and Low Merge peaks and valleys data points using pandas extrema
        df_highs_peaks = pd.DataFrame({'Date': h.index[highs_peaks], 'Extremum': h.values[highs_peaks]})
        df_lows_valleys = pd.DataFrame({'Date': l.index[lows_valleys], 'Extremum': l.values[lows_valleys]})

        ex1_df_highs_peaks_and_lows_valleys = pd.concat([df_highs_peaks, df_lows_valleys], axis=0, ignore_index=True, sort=True)
        ex1_df_highs_peaks_and_lows_valleys = ex1_df_highs_peaks_and_lows_valleys.sort_values(by=['Date'])

        # Apply the ZigZag filter
        ex1_extrema = []
        ex1_extrema.append((last_extrema_2_date, last_extrema_2_value))

        for i in range(len(ex1_df_highs_peaks_and_lows_valleys)):
            if (((ex1_df_highs_peaks_and_lows_valleys.iloc[i][['Date']] == df_highs_peaks[['Date']]).all(axis=1).any()) and ((ex1_extrema[-1][0] == df_highs_peaks[['Date']]).all(axis=1).any())):
                continue

            if (((ex1_df_highs_peaks_and_lows_valleys.iloc[i][['Date']] == df_lows_valleys[['Date']]).all(axis=1).any()) and ((ex1_extrema[-1][0] == df_lows_valleys[['Date']]).all(axis=1).any())):
                continue

            else:
                ex1_extrema.append((ex1_df_highs_peaks_and_lows_valleys['Date'].iloc[i], ex1_df_highs_peaks_and_lows_valleys['Extremum'].iloc[i]))


        # Convert ex1_extrema to DataFrame
        ex1_df_extrema = pd.DataFrame(ex1_extrema, columns=['Date', 'Extremum'])
        ex1_df_extrema = ex1_df_extrema.sort_values(by='Date')



        #EXTREMA 2
        # Convert the 'date' column to a list of datetime objects
        df_list_peaks_dates = df_highs_peaks['Date'].tolist()
        df_list_valleys_dates = df_lows_valleys['Date'].tolist()

        # Create a new DataFrame to store the new data
        ex1_df_highs_peaks = []
        ex1_df_lows_valleys = []

        for date, extremum in ex1_extrema:
            if date in df_list_peaks_dates:
                ex1_df_highs_peaks.append({'Date': date, 'Extremum': extremum})


        for date, extremum in ex1_extrema:
            if date in df_list_valleys_dates:
                ex1_df_lows_valleys.append({'Date': date, 'Extremum': extremum})


        # Convert the ex2_df_highs_peaks list to a DataFrame
        ex1_df_highs_peaks = pd.DataFrame(ex1_df_highs_peaks)
        ex1_df_lows_valleys = pd.DataFrame(ex1_df_lows_valleys)

        # Create a new DataFrame to store the new data
        ex2_df_highs_peaks = []
        ex2_df_lows_valleys = []

        # Iterate through ex1_extrema and add new data to the ex2_df_highs_peaks list
        for date, extremum in ex1_extrema:
            if date in df_list_peaks_dates:
                ex2_df_highs_peaks.append({'Date': date, 'Extremum': extremum})


        for date, extremum in ex1_extrema:
            if date in df_list_valleys_dates:
                ex2_df_lows_valleys.append({'Date': date, 'Extremum': extremum})


        # Convert the ex2_df_highs_peaks list to a DataFrame
        ex2_df_highs_peaks = pd.DataFrame(ex2_df_highs_peaks)
        ex2_df_lows_valleys = pd.DataFrame(ex2_df_lows_valleys)

        # Find local extrema using argrelextrema
        ex2_extrema_highs_peaks = signal.argrelextrema(ex2_df_highs_peaks['Extremum'].values, np.greater, order=1)
        ex2_extrema_lows_valleys = signal.argrelextrema(ex2_df_lows_valleys['Extremum'].values, np.less, order=1)

        # Retrieve dates and values
        ex2_highs_peaks_dates = ex2_df_highs_peaks.loc[ex2_extrema_highs_peaks[0], 'Date']
        ex2_highs_peaks_values = ex2_df_highs_peaks.loc[ex2_extrema_highs_peaks[0], 'Extremum']

        ex2_lows_valleys_dates = ex2_df_lows_valleys.loc[ex2_extrema_lows_valleys[0], 'Date']
        ex2_lows_valleys_values = ex2_df_lows_valleys.loc[ex2_extrema_lows_valleys[0], 'Extremum']

        # Create a DataFrame with extrema information
        ex2_df_extrema_highs_peaks = pd.DataFrame({'Date': ex2_highs_peaks_dates, 'Extremum': ex2_highs_peaks_values})
        ex2_df_extrema_lows_valleys = pd.DataFrame({'Date': ex2_lows_valleys_dates, 'Extremum': ex2_lows_valleys_values})   

        # Concatenate the two dataframes
        ex2_df_highs_peaks_and_lows_valleys = pd.concat([ex2_df_extrema_highs_peaks, ex2_df_extrema_lows_valleys], ignore_index=True)
        ex2_df_highs_peaks_and_lows_valleys = pd.DataFrame(ex2_df_highs_peaks_and_lows_valleys, columns=['Date', 'Extremum'])
        ex2_df_highs_peaks_and_lows_valleys = ex2_df_highs_peaks_and_lows_valleys.sort_values(by='Date')

        #Filter for extrema
        ex2_extrema = []
        ex2_extrema.append((last_extrema_2_date, last_extrema_2_value))

        for i in range(len(ex2_df_highs_peaks_and_lows_valleys)):
            if (((ex2_df_highs_peaks_and_lows_valleys.iloc[i][['Date']] == df_highs_peaks[['Date']]).all(axis=1).any()) and ((ex2_extrema[-1][0] == df_highs_peaks[['Date']]).all(axis=1).any())):
                continue

            if (((ex2_df_highs_peaks_and_lows_valleys.iloc[i][['Date']] == df_lows_valleys[['Date']]).all(axis=1).any()) and ((ex2_extrema[-1][0] == df_lows_valleys[['Date']]).all(axis=1).any())):
                continue

            else:
                ex2_extrema.append((ex2_df_highs_peaks_and_lows_valleys['Date'].iloc[i], ex2_df_highs_peaks_and_lows_valleys['Extremum'].iloc[i]))


        # Convert ex2_df_extrema to DataFrame
        ex2_df_extrema = pd.DataFrame(ex2_extrema, columns=['Date', 'Extremum'])
        ex2_df_extrema = ex2_df_extrema.sort_values(by='Date')

        data = data[last_date_info < data.index]
        ex1_df_extrema = ex1_df_extrema[last_date_info < ex1_df_extrema['Date']]
        ex2_df_extrema = ex2_df_extrema[last_date_info < ex2_df_extrema['Date']]

        return data, ex1_df_extrema, ex2_df_extrema

    except Exception as e:
        return False



async def mysql_update_new_data(symbol):
    # Iterate over the rows in the filtered DataFrame and update the MySQL table
    for index, row in data.iterrows():
        open_val = row['Open']
        high_val = row['High']
        low_val = row['Low']
        close_val = row['Close']

        cursor.execute(f'''
            INSERT INTO `{symbol}` (Date_Info, Open, High, Low, Close)
            VALUES (%s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                Open = VALUES(Open),
                High = VALUES(High),
                Low = VALUES(Low),
                Close = VALUES(Close)
        ''', (index, open_val, high_val, low_val, close_val))

    if not ex1_df_extrema.empty:
        for _, row in ex1_df_extrema.iterrows():
            cursor.execute(f'''
                UPDATE `{symbol}`
                SET Extrema_1 = %s
                WHERE Date_Info = %s;
            ''', (row['Extremum'], row['Date']))

    if not ex2_df_extrema.empty:
        for _, row in ex2_df_extrema.iterrows():
            cursor.execute(f'''
                UPDATE `{symbol}`
                SET Extrema_2 = %s
                WHERE Date_Info = %s;
            ''', (row['Extremum'], row['Date']))

    conn.commit()



async def mysql_close_connection(cursor, conn):
    cursor.close()
    conn.close()



async def get_market_structure_async(symbol, cursor):
    try:
        market_structure_df = pd.DataFrame(columns=['Symbol', 'Market Structure'], index=None)

        cursor.execute(f'''SELECT Date_Info FROM `{symbol}` WHERE Date_Info IS NOT NULL ORDER BY Date_Info DESC LIMIT 1 OFFSET 1''')
        second_last_date_info = cursor.fetchone()[0]

        cursor.execute(f'''SELECT Date_Info FROM `{symbol}` WHERE Date_Info IS NOT NULL ORDER BY Date_Info DESC LIMIT 1''')
        last_date_info = cursor.fetchone()[0]

        cursor.execute(f'''SELECT Date_Info FROM `{symbol}` WHERE Extrema_1 IS NOT NULL ORDER BY Date_Info DESC LIMIT 1 OFFSET 2''')
        third_last_extrema_1_date = cursor.fetchone()[0]

        cursor.execute(f'''SELECT Date_Info FROM `{symbol}` WHERE Extrema_1 IS NOT NULL ORDER BY Date_Info DESC LIMIT 1''')
        last_extrema_1_date = cursor.fetchone()[0]

        cursor.execute(f'''SELECT Extrema_1 FROM `{symbol}` WHERE Extrema_1 IS NOT NULL ORDER BY Date_Info DESC LIMIT 1''')
        last_extrema_1_value = cursor.fetchone()[0]

        cursor.execute(f'''SELECT Date_Info FROM `{symbol}` WHERE Extrema_2 IS NOT NULL ORDER BY Date_Info DESC LIMIT 1''')
        last_extrema_2_date = cursor.fetchone()[0]

        cursor.execute(f'''SELECT Extrema_2 FROM `{symbol}` WHERE Extrema_2 IS NOT NULL ORDER BY Date_Info DESC LIMIT 1''')
        last_extrema_2_value = cursor.fetchone()[0]

        Market_Structure = 'None'

        if ((second_last_date_info == last_extrema_1_date) and
            (third_last_extrema_1_date == last_extrema_2_date) and
            (last_extrema_2_value < last_extrema_1_value)):
            Market_Structure = 'Expected Bull Trend'

        if ((second_last_date_info == last_extrema_1_date) and
            (third_last_extrema_1_date == last_extrema_2_date) and
            (last_extrema_2_value > last_extrema_1_value)):
            Market_Structure = 'Expected Bear Trend'

        if (Market_Structure == 'None'):
            return False

        if (Market_Structure != 'None'):
            print("Symbol:", symbol)
            print('Market Structure: ', Market_Structure)
            market_structure_df = pd.concat([market_structure_df, pd.DataFrame({'Symbol': [symbol], 'Market Structure': [Market_Structure]})], ignore_index=True)
            return market_structure_df

    except Exception as e:
        return False



async def get_trailing_pe_async(session, symbol):
    try:
        trailing_pe_df = pd.DataFrame(columns=['Symbol', 'Trailing P/E'], index=None)
        field = 'Trailing P/E'  # The field we want to extract

        url = f'https://finance.yahoo.com/quote/{symbol}/key-statistics?p={symbol}'
        headers = {'User-agent': 'Mozilla/5.0'}

        async with session.get(url, headers=headers) as response:
            html_content = await response.text()

        soup = BeautifulSoup(html_content, features='lxml')
        tables = soup.findAll('table')
        trailing_pe = None

        for table in tables:
            for tr in table.find_all('tr'):
                row = [td.text.strip() for td in tr.find_all('td')]
                try:
                    if row[0] == field and row[1] != 'N/A':
                        trailing_pe = row[1]
                        break
                except AttributeError:
                    # Handle missing data gracefully
                    print(f'{row[0]} data not found for {symbol}')

        trailing_pe_df = pd.concat([trailing_pe_df, pd.DataFrame({'Symbol': [symbol], 'Trailing P/E': [trailing_pe]})], ignore_index=True)
        print('Symbol:', symbol)
        print('P/E:', trailing_pe, '\n')
        return trailing_pe_df

    except AttributeError:
        return False



async def get_enterprise_value_ebitda_async(session, symbol):
    try:
        enterprise_value_ebitda_df = pd.DataFrame(columns=['Symbol', 'EV/EBITDA'], index=None)
        field = 'Enterprise Value/EBITDA'  # The field we want to extract

        url = f'https://finance.yahoo.com/quote/{symbol}/key-statistics?p={symbol}'
        headers = {'User-agent': 'Mozilla/5.0'}

        async with session.get(url, headers=headers) as response:
            html_content = await response.text()

        soup = BeautifulSoup(html_content, features='lxml')
        tables = soup.findAll('table')
        enterprise_value_ebitda = None

        for table in tables:
            for tr in table.find_all('tr'):
                row = [td.text.strip() for td in tr.find_all('td')]
                try:
                    if row[0] == field and row[1] != 'N/A':
                        enterprise_value_ebitda = row[1]
                        break
                except AttributeError:
                    # Handle missing data gracefully
                    print(f'{row[0]} data not found for {symbol}')

        enterprise_value_ebitda_df = pd.concat([enterprise_value_ebitda_df, pd.DataFrame({'Symbol': [symbol], 'EV/EBITDA': [enterprise_value_ebitda]})], ignore_index=True)
        print('Symbol:', symbol)
        print('EV/EBITDA:', enterprise_value_ebitda, '\n')
        return enterprise_value_ebitda_df

    except AttributeError:
        return False



async def get_price_sales_async(session, symbol):
    try:
        price_sales_df = pd.DataFrame(columns=['Symbol', 'Price/Sales (ttm)'], index=None)
        field = 'Price/Sales (ttm)'  # The field we want to extract

        url = f'https://finance.yahoo.com/quote/{symbol}/key-statistics?p={symbol}'
        headers = {'User-agent': 'Mozilla/5.0'}

        async with session.get(url, headers=headers) as response:
            html_content = await response.text()

        soup = BeautifulSoup(html_content, features='lxml')
        tables = soup.findAll('table')
        price_sales = None

        for table in tables:
            for tr in table.find_all('tr'):
                row = [td.text.strip() for td in tr.find_all('td')]
                try:
                    if row[0] == field and row[1] != 'N/A':
                        price_sales = row[1]
                        break
                except AttributeError:
                    # Handle missing data gracefully
                    print(f'{row[0]} data not found for {symbol}')

        price_sales_df = pd.concat([price_sales_df, pd.DataFrame({'Symbol': [symbol], 'Price/Sales (ttm)': [price_sales]})], ignore_index=True)
        print('Symbol:', symbol)
        print('P/S:', price_sales, '\n')
        return price_sales_df

    except AttributeError:
        return False



async def get_price_book_async(session, symbol):
    try:
        price_book_df = pd.DataFrame(columns=['Symbol', 'Price/Book (mrq)'], index=None)
        field = 'Price/Book (mrq)'  # The field we want to extract

        url = f'https://finance.yahoo.com/quote/{symbol}/key-statistics?p={symbol}'
        headers = {'User-agent': 'Mozilla/5.0'}

        async with session.get(url, headers=headers) as response:
            html_content = await response.text()

        soup = BeautifulSoup(html_content, features='lxml')
        tables = soup.findAll('table')
        price_book = None

        for table in tables:
            for tr in table.find_all('tr'):
                row = [td.text.strip() for td in tr.find_all('td')]
                try:
                    if row[0] == field and row[1] != 'N/A':
                        price_book = row[1]
                        break
                except AttributeError:
                    # Handle missing data gracefully
                    print(f'{row[0]} data not found for {symbol}')

        price_book_df = pd.concat([price_book_df, pd.DataFrame({'Symbol': [symbol], 'Price/Book (mrq)': [price_book]})], ignore_index=True)
        print('Symbol:', symbol)
        print('P/B:', price_book, '\n')
        return price_book_df

    except AttributeError:
        return False



async def get_profit_margin_async(session, symbol):
    try:
        profit_margin_df = pd.DataFrame(columns=['Symbol', 'Profit Margin'], index=None)
        field = 'Profit Margin'  # The field we want to extract

        url = f'https://finance.yahoo.com/quote/{symbol}/key-statistics?p={symbol}'
        headers = {'User-agent': 'Mozilla/5.0'}

        async with session.get(url, headers=headers) as response:
            html_content = await response.text()

        soup = BeautifulSoup(html_content, features='lxml')
        tables = soup.findAll('table')
        profit_margin = None

        for table in tables:
            for tr in table.find_all('tr'):
                row = [td.text.strip() for td in tr.find_all('td')]
                try:
                    if row[0] == field and row[1] != 'N/A':
                        profit_margin = row[1]
                        break
                except AttributeError:
                    # Handle missing data gracefully
                    print(f'{row[0]} data not found for {symbol}')

        profit_margin_df = pd.concat([profit_margin_df, pd.DataFrame({'Symbol': [symbol], 'Profit Margin': [profit_margin]})], ignore_index=True)
        print('Symbol:', symbol)
        print('Profit Margin:', profit_margin, '\n')
        return profit_margin_df

    except AttributeError:
        return False



async def get_return_on_equity_async(session, symbol):
    try:
        return_on_equity_df = pd.DataFrame(columns=['Symbol', 'ROE'], index=None)
        field = 'Return on Equity (ttm)'  # The field we want to extract

        url = f'https://finance.yahoo.com/quote/{symbol}/key-statistics?p={symbol}'
        headers = {'User-agent': 'Mozilla/5.0'}

        async with session.get(url, headers=headers) as response:
            html_content = await response.text()

        soup = BeautifulSoup(html_content, features='lxml')
        tables = soup.findAll('table')
        return_on_equity = None

        for table in tables:
            for tr in table.find_all('tr'):
                row = [td.text.strip() for td in tr.find_all('td')]
                try:
                    if row[0] == field and row[1] != 'N/A':
                        return_on_equity = row[1]
                        break
                except AttributeError:
                    # Handle missing data gracefully
                    print(f'{row[0]} data not found for {symbol}')

        return_on_equity_df = pd.concat([return_on_equity_df, pd.DataFrame({'Symbol': [symbol], 'ROE': [return_on_equity]})], ignore_index=True)
        print('Symbol:', symbol)
        print('ROE:', return_on_equity, '\n')
        return return_on_equity_df

    except AttributeError:
        return False



async def get_return_on_assets_async(session, symbol):
    try:
        return_on_assets_df = pd.DataFrame(columns=['Symbol', 'ROA'], index=None)
        field = 'Return on Assets (ttm)'  # The field we want to extract

        url = f'https://finance.yahoo.com/quote/{symbol}/key-statistics?p={symbol}'
        headers = {'User-agent': 'Mozilla/5.0'}

        async with session.get(url, headers=headers) as response:
            html_content = await response.text()

        soup = BeautifulSoup(html_content, features='lxml')
        tables = soup.findAll('table')
        return_on_assets = None

        for table in tables:
            for tr in table.find_all('tr'):
                row = [td.text.strip() for td in tr.find_all('td')]
                try:
                    if row[0] == field and row[1] != 'N/A':
                        return_on_assets = row[1]
                        break
                except AttributeError:
                    # Handle missing data gracefully
                    print(f'{row[0]} data not found for {symbol}')

        return_on_assets_df = pd.concat([return_on_assets_df, pd.DataFrame({'Symbol': [symbol], 'ROA': [return_on_assets]})], ignore_index=True)
        print('Symbol:', symbol)
        print('ROA:', return_on_assets, '\n')
        return return_on_assets_df

    except AttributeError:
        return False



async def get_current_ratio_async(session, symbol):
    try:
        current_ratio_df = pd.DataFrame(columns=['Symbol', 'Current Ratio (mrq)'], index=None)
        field = 'Current Ratio (mrq)'  # The field we want to extract

        url = f'https://finance.yahoo.com/quote/{symbol}/key-statistics?p={symbol}'
        headers = {'User-agent': 'Mozilla/5.0'}

        async with session.get(url, headers=headers) as response:
            html_content = await response.text()

        soup = BeautifulSoup(html_content, features='lxml')
        tables = soup.findAll('table')
        current_ratio = None

        for table in tables:
            for tr in table.find_all('tr'):
                row = [td.text.strip() for td in tr.find_all('td')]
                try:
                    if row[0] == field and row[1] != 'N/A':
                        current_ratio = row[1]
                        break
                except AttributeError:
                    # Handle missing data gracefully
                    print(f'{row[0]} data not found for {symbol}')

        current_ratio_df = pd.concat([current_ratio_df, pd.DataFrame({'Symbol': [symbol], 'Current Ratio (mrq)': [current_ratio]})], ignore_index=True)
        print('Symbol:', symbol)
        print('Current Ratio (mrq):', current_ratio, '\n')
        return current_ratio_df

    except AttributeError:
        return False



async def get_total_debt_equity_async(session, symbol):
    try:
        total_debt_equity_df = pd.DataFrame(columns=['Symbol', 'Total Debt/Equity (mrq)'], index=None)
        field = 'Total Debt/Equity (mrq)'  # The field we want to extract

        url = f'https://finance.yahoo.com/quote/{symbol}/key-statistics?p={symbol}'
        headers = {'User-agent': 'Mozilla/5.0'}

        async with session.get(url, headers=headers) as response:
            html_content = await response.text()

        soup = BeautifulSoup(html_content, features='lxml')
        tables = soup.findAll('table')
        total_debt_equity = None

        for table in tables:
            for tr in table.find_all('tr'):
                row = [td.text.strip() for td in tr.find_all('td')]
                try:
                    if row[0] == field and row[1] != 'N/A':
                        total_debt_equity = row[1]
                        break
                except AttributeError:
                    # Handle missing data gracefully
                    print(f'{row[0]} data not found for {symbol}')

        total_debt_equity_df = pd.concat([total_debt_equity_df, pd.DataFrame({'Symbol': [symbol], 'Total Debt/Equity (mrq)': [total_debt_equity]})], ignore_index=True)
        print('Symbol:', symbol)
        print('Total Debt/Equity (mrq):', total_debt_equity, '\n')
        return total_debt_equity_df

    except AttributeError:
        return False



async def get_trailing_dividend_yield_async(session, symbol):
    try:
        trailing_dividend_yield_df = pd.DataFrame(columns=['Symbol', 'Trailing Dividend Yield'], index=None)
        field_prefix = 'Trailing Annual Dividend Yield'  # The field prefix we want to extract

        url = f'https://finance.yahoo.com/quote/{symbol}/key-statistics?p={symbol}'
        headers = {'User-agent': 'Mozilla/5.0'}

        async with session.get(url, headers=headers) as response:
            html_content = await response.text()

        soup = BeautifulSoup(html_content, features='lxml')
        tables = soup.findAll('table')
        trailing_dividend_yield = None

        for table in tables:
            for tr in table.find_all('tr'):
                row = [td.text.strip() for td in tr.find_all('td')]
                try:
                    if row[0].startswith(field_prefix) and row[1] != 'N/A':
                        trailing_dividend_yield = row[1]
                        trailing_dividend_yield_df = pd.concat([trailing_dividend_yield_df, pd.DataFrame({'Symbol': [symbol], 'Trailing Dividend Yield': [trailing_dividend_yield]})], ignore_index=True)
                except AttributeError:
                    # Handle missing data gracefully
                    print(f'{row[0]} data not found for {symbol}')

        
        print('Symbol:', symbol)
        print('Trailing Dividend Yield:', trailing_dividend_yield, '\n')
        return trailing_dividend_yield_df

    except AttributeError:
        return False



async def get_payout_ratio_async(session, symbol):
    try:
        payout_ratio_df = pd.DataFrame(columns=['Symbol', 'Payout Ratio'], index=None)
        field_prefix = 'Payout Ratio'  # The field prefix we want to match

        url = f'https://finance.yahoo.com/quote/{symbol}/key-statistics?p={symbol}'
        headers = {'User-agent': 'Mozilla/5.0'}

        async with session.get(url, headers=headers) as response:
            html_content = await response.text()

        soup = BeautifulSoup(html_content, features='lxml')
        tables = soup.findAll('table')
        payout_ratio = None

        for table in tables:
            for tr in table.find_all('tr'):
                row = [td.text.strip() for td in tr.find_all('td')]
                try:
                    if row[0].startswith(field_prefix) and row[1] != 'N/A':
                        payout_ratio = row[1]
                        payout_ratio_df = pd.concat([payout_ratio_df, pd.DataFrame({'Symbol': [symbol], 'Payout Ratio': [payout_ratio]})], ignore_index=True)
                except AttributeError:
                    # Handle missing data gracefully
                    print(f'{row[0]} data not found for {symbol}')

        print('Symbol:', symbol)
        print('Payout Ratio:', payout_ratio, '\n')
        return payout_ratio_df

    except AttributeError:
        return False



# Separate function to filter symbols and call respective functions
async def process_symbols(symbols, batch_number):
    async with aiohttp.ClientSession() as session:
        tasks = [get_market_cap_async(session, symbol) for symbol in symbols]
        results = await asyncio.gather(*tasks)
        data = []
        tasks = []

        for i, symbol in enumerate(symbols):
            if results[i] is not False:
                market_cap_df = results[i]
                conn, cursor, table_exists = await mysql_databse_connection(symbol)

                #Not Existing Table
                if not table_exists:

                    result = await extrema(symbol)
                    if result is False:
                        continue

                    ex1_df_extrema, ex2_df_extrema, data = await extrema(symbol)
                    await mysql_not_existing_table(symbol, conn, cursor, data, ex1_df_extrema, ex2_df_extrema)
                    market_structure_result = await get_market_structure_async(symbol, cursor)

                #Existing Table
                if table_exists:
                    last_date_info = await mysql_existing_table_checking_new_data(symbol, cursor)
                    result = await checking_new_data(symbol, last_date_info)

                    #There is Not New Data
                    if result == False:
                        continue

                    #There is New Data
                    last_extrema_1_value, last_extrema_1_date, last_extrema_2_value, last_extrema_2_date = await mysql_last_data(symbol, cursor)

                    result = await new_extrema(symbol)
                    if result is False:
                        continue

                    data, ex1_df_extrema, ex2_df_extrema = await new_extrema(symbol)
                    await mysql_update_new_data(symbol)
                    market_structure_result = await get_market_structure_async(symbol, cursor)

                if market_structure_result is not False:
                    market_structure_df = market_structure_result
                    result = await get_sector_async(session, symbol)
                    if result is not False:
                        sector_df, industry_df = await get_sector_async(session, symbol)
                        trailing_pe_df = await get_trailing_pe_async(session, symbol)
                        enterprise_value_ebitda_df = await get_enterprise_value_ebitda_async(session, symbol)
                        price_sales_df = await get_price_sales_async(session, symbol)
                        price_book_df = await get_price_book_async(session, symbol)
                        profit_margin_df = await get_profit_margin_async(session, symbol)
                        return_on_equity_df = await get_return_on_equity_async(session, symbol)
                        return_on_assets_df = await get_return_on_assets_async(session, symbol)
                        current_ratio_df = await get_current_ratio_async(session, symbol)
                        total_debt_equity_df = await get_total_debt_equity_async(session, symbol)
                        trailing_dividend_yield_df = await get_trailing_dividend_yield_async(session, symbol)
                        payout_ratio_df = await get_payout_ratio_async(session, symbol)
                        
                        sector_data.append(sector_df)
                        industry_data.append(industry_df)
                        market_cap_data.append(market_cap_df)
                        market_structure_data.append(market_structure_df)
                        trailing_pe_data.append(trailing_pe_df)
                        enterprise_value_ebitda_data.append(enterprise_value_ebitda_df)
                        price_sales_data.append(price_sales_df)
                        price_book_data.append(price_book_df)
                        profit_margin_data.append(profit_margin_df)
                        return_on_equity_data.append(return_on_equity_df)
                        return_on_assets_data.append(return_on_assets_df)
                        current_ratio_data.append(current_ratio_df)
                        total_debt_equity_data.append(total_debt_equity_df)
                        trailing_dividend_yield_data.append(trailing_dividend_yield_df)
                        payout_ratio_data.append(payout_ratio_df)
                
            else:
                continue
              
        await asyncio.gather(*tasks)
        # await mysql_close_connection(cursor, conn)


if __name__ == "__main__":
    df_nasdaq_symbols = df_nasdaq['Symbol'].tolist()
    df_other_symbols = df_other['Symbol'].tolist()
    all_us_equity_df = df_nasdaq_symbols + df_other_symbols
    # all_us_equity_df = all_us_equity_df[4650:]
    print('len all_us_equity_df:\n', len(all_us_equity_df))

    sector_data = []
    industry_data = []
    market_cap_data = []
    market_structure_data = []
    trailing_pe_data = []
    enterprise_value_ebitda_data = []
    price_sales_data = []
    price_book_data = []
    profit_margin_data = []
    return_on_equity_data = []
    return_on_assets_data = []
    current_ratio_data = []
    total_debt_equity_data = []
    trailing_dividend_yield_data = []
    payout_ratio_data = []
    
    # Adjust the batch size as per your preference
    batch_size = 10
    symbol_batches = [all_us_equity_df[i:i + batch_size] for i in range(0, len(all_us_equity_df), batch_size)]

    # Define a counter for the number of symbols processed
    symbols_processed = 0
    batch_number = 1

    # Run the tasks asynchronously in batches
    for batch in symbol_batches:
        print('Index: ', symbols_processed, '/', len(all_us_equity_df))
        if symbols_processed >= len(all_us_equity_df):
            break  # Stop processing after 200 symbols

        remaining_symbols = len(all_us_equity_df) - symbols_processed
        current_batch_size = min(len(batch), remaining_symbols)

        asyncio.run(process_symbols(batch[:current_batch_size], batch_number))

        symbols_processed += current_batch_size
        batch_number += 1


    sector_data = pd.concat(sector_data)
    sector_data.set_index('Symbol', inplace=True)

    industry_data = pd.concat(industry_data)
    industry_data.set_index('Symbol', inplace=True)

    market_cap_data = pd.concat(market_cap_data)
    market_cap_data.set_index('Symbol', inplace=True)

    market_structure_data = pd.concat(market_structure_data)
    market_structure_data.set_index('Symbol', inplace=True)

    trailing_pe_data = pd.concat(trailing_pe_data)
    trailing_pe_data.set_index('Symbol', inplace=True)

    enterprise_value_ebitda_data = pd.concat(enterprise_value_ebitda_data)
    enterprise_value_ebitda_data.set_index('Symbol', inplace=True)

    price_sales_data = pd.concat(price_sales_data)
    price_sales_data.set_index('Symbol', inplace=True)

    price_book_data = pd.concat(price_book_data)
    price_book_data.set_index('Symbol', inplace=True)

    profit_margin_data = pd.concat(profit_margin_data)
    profit_margin_data.set_index('Symbol', inplace=True)

    return_on_equity_data = pd.concat(return_on_equity_data)
    return_on_equity_data.set_index('Symbol', inplace=True)

    return_on_assets_data = pd.concat(return_on_assets_data)
    return_on_assets_data.set_index('Symbol', inplace=True)

    current_ratio_data = pd.concat(current_ratio_data)
    current_ratio_data.set_index('Symbol', inplace=True)

    total_debt_equity_data = pd.concat(total_debt_equity_data)
    total_debt_equity_data.set_index('Symbol', inplace=True)

    trailing_dividend_yield_data = pd.concat(trailing_dividend_yield_data)
    trailing_dividend_yield_data.set_index('Symbol', inplace=True)

    payout_ratio_data = pd.concat(payout_ratio_data)
    payout_ratio_data.set_index('Symbol', inplace=True)

    # data_frames = [sector_data, market_cap_data, price_action_data, profit_margin_data, return_on_equity_data, return_on_assets_data, trailing_pe_data, price_sales_data, price_book_data, peg_data, enterprise_value_ebitda_data, total_debt_equity_data, current_ratio_data, payout_ratio_data, short_float_data]
    combined_data = pd.concat([sector_data, industry_data, market_cap_data, market_structure_data, trailing_pe_data, enterprise_value_ebitda_data, price_sales_data, price_book_data, profit_margin_data, return_on_equity_data, return_on_assets_data, current_ratio_data, total_debt_equity_data, trailing_dividend_yield_data, payout_ratio_data], axis=1)
    combined_data_sorted = combined_data.sort_values(by='Symbol')

    # Create an Excel writer object and set the alignment
    current_datetime = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    excel_file = f'Tangra_Stock_Financial_Analyst_{current_datetime}.xlsx'
    excel_writer = pd.ExcelWriter(excel_file, engine='openpyxl')
    combined_data_sorted.to_excel(excel_writer, index=True, sheet_name='Sheet1', startrow=2)

    # Get the workbook and the worksheet
    workbook = excel_writer.book
    worksheet = workbook['Sheet1']
    
    worksheet.insert_cols(5)
    worksheet.insert_cols(7)
    worksheet.insert_cols(12)
    worksheet.insert_cols(16)
    worksheet.insert_cols(19)

    # Define the alignment setting
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')


    for col_idx, col in enumerate(worksheet.columns):
        if col_idx == 0:
            for cell in col:
                cell.alignment = align_center


    # Iterate through columns to set alignment
    for col_idx, col in enumerate(worksheet.columns):
        if col_idx == 0:  # Skip the index column
            continue
        for cell in col:
            cell.alignment = align_right


    price_action_column = 'F'

    # Initialize a row index variable
    row_index = 4

    for row_index, row in enumerate(worksheet.iter_rows(min_row=4, values_only=True), start=4):
        value = row[5]
        if isinstance(value, str):
            if value.startswith('Expected Bull'):
                worksheet[f'{price_action_column}{row_index}'].fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green background
            elif value.startswith('Expected Bear'):
                worksheet[f'{price_action_column}{row_index}'].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red background


    trailing_pe_column = 'H'

    for row_index, row in enumerate(worksheet.iter_rows(min_row=4, values_only=True), start=4):
        value = row[7]

        # Convert to float if not already
        if isinstance(value, str):
            match = re.search(r'-?\d+(\.\d*)?', value)
            if match:
                value = float(match.group())
            else:
                value = None  # Handle invalid cases

        if value is not None:
            if value < 20:
                worksheet[f'{trailing_pe_column}{row_index}'].fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green background
            elif value > 25:
                worksheet[f'{trailing_pe_column}{row_index}'].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red background


    enterprise_value_ebitda_column = 'I'

    for row_index, row in enumerate(worksheet.iter_rows(min_row=4, values_only=True), start=4):
        value = row[8]

        # Convert to float if not already
        if isinstance(value, str):
            match = re.search(r'-?\d+(\.\d*)?', value)
            if match:
                value = float(match.group())
            else:
                value = None  # Handle invalid cases

        if value is not None:
            # if value < 1:
            #     worksheet[f'{enterprise_value_ebitda_column}{row_index}'].fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green background
            if value < 0:
                worksheet[f'{enterprise_value_ebitda_column}{row_index}'].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red background


    price_book_column = 'K'

    for row_index, row in enumerate(worksheet.iter_rows(min_row=4, values_only=True), start=4):
        value = row[10]

        # Convert to float if not already
        if isinstance(value, str):
            match = re.search(r'-?\d+(\.\d*)?', value)
            if match:
                value = float(match.group())
            else:
                value = None  # Handle invalid cases

        if value is not None:
            if value < 1:
                worksheet[f'{price_book_column}{row_index}'].fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green background
            elif value > 3:
                worksheet[f'{price_book_column}{row_index}'].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red background
    

    profit_margin_column = 'M'

    for row_index, row in enumerate(worksheet.iter_rows(min_row=4, values_only=True), start=4):
        value = row[12]

        # Convert to float if not already
        if isinstance(value, str):
            match = re.search(r'-?\d+(\.\d*)?', value)
            if match:
                value = float(match.group())
            else:
                value = None  # Handle invalid cases

        if value is not None:
            if value >= 0:
                worksheet[f'{profit_margin_column}{row_index}'].fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green background
            elif value < 0:
                worksheet[f'{profit_margin_column}{row_index}'].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red background
    

    return_on_equity_column = 'N'

    for row_index, row in enumerate(worksheet.iter_rows(min_row=4, values_only=True), start=4):
        value = row[13]

        # Convert to float if not already
        if isinstance(value, str):
            match = re.search(r'-?\d+(\.\d*)?', value)
            if match:
                value = float(match.group())
            else:
                value = None  # Handle invalid cases

        if value is not None:
            if value >=10:
                worksheet[f'{return_on_equity_column}{row_index}'].fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green background
            elif value < 10:
                worksheet[f'{return_on_equity_column}{row_index}'].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red background


    return_on_assets_column = 'O' 

    for row_index, row in enumerate(worksheet.iter_rows(min_row=4, values_only=True), start=4):
        value = row[13]

        # Convert to float if not already
        if isinstance(value, str):
            match = re.search(r'-?\d+(\.\d*)?', value)
            if match:
                value = float(match.group())
            else:
                value = None  # Handle invalid cases

        if value is not None:
            # if value >=10:
            #     worksheet[f'{return_on_assets_column}{row_index}'].fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green background
            if value < 0:
                worksheet[f'{return_on_assets_column}{row_index}'].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red background


    current_ratio_column = 'Q'

    for row_index, row in enumerate(worksheet.iter_rows(min_row=4, values_only=True), start=4):
        value = row[16]

        # Convert to float if not already
        if isinstance(value, str):
            match = re.search(r'-?\d+(\.\d*)?', value)
            if match:
                value = float(match.group())
            else:
                value = None  # Handle invalid cases

        if value is not None:
            if value >= 1:
                worksheet[f'{current_ratio_column}{row_index}'].fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green background
            elif value < 1:
                worksheet[f'{current_ratio_column}{row_index}'].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red background


    total_debt_equity_column = 'R'

    for row_index, row in enumerate(worksheet.iter_rows(min_row=4, values_only=True), start=4):
        value = row[17]

        # Convert to float if not already
        if isinstance(value, str):
            match = re.search(r'-?\d+(\.\d*)?', value)
            if match:
                value = float(match.group())
            else:
                value = None  # Handle invalid cases

        if value is not None:
            if value <= 1:
                worksheet[f'{total_debt_equity_column}{row_index}'].fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green background
            elif value > 1:
                worksheet[f'{total_debt_equity_column}{row_index}'].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red background


    # Increment the row index after processing each row
    row_index += 1


    # Apply bold font style
    bold_font = Font(bold=True)
    blue_fill = PatternFill(start_color='33CCCC', end_color='33CCCC', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')


    worksheet.merge_cells('A1:D1')
    company_profile_cell = worksheet['A1']
    company_profile_cell.value = "Company Profile"
    company_profile_cell.font = bold_font
    company_profile_cell.fill = blue_fill
    company_profile_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    technical_analysis_cell = worksheet['F1']
    technical_analysis_cell.value = 'Technical Analysis'
    technical_analysis_cell.font = bold_font
    technical_analysis_cell.fill = blue_fill
    technical_analysis_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    worksheet.merge_cells('H1:U1')
    fundamental_analysis_cell = worksheet['H1']
    fundamental_analysis_cell.value = 'Fundamental Analysis'
    fundamental_analysis_cell.font = bold_font
    fundamental_analysis_cell.fill = blue_fill
    fundamental_analysis_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')


    worksheet.merge_cells('H2:K2')
    key_ratios_cell = worksheet['H2']
    key_ratios_cell.value = 'Key Ratios'
    key_ratios_cell.font = bold_font
    key_ratios_cell.fill = yellow_fill
    key_ratios_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    worksheet.merge_cells('M2:O2')
    profitability_cell = worksheet['M2']
    profitability_cell.value = 'Profitability (ttm)'
    profitability_cell.font = bold_font
    profitability_cell.fill = yellow_fill
    profitability_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    worksheet.merge_cells('Q2:R2')
    liquidity_cell = worksheet['Q2']
    liquidity_cell.value = 'Liquidity (mrq)'
    liquidity_cell.font = bold_font
    liquidity_cell.fill = yellow_fill
    liquidity_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    worksheet.merge_cells('T2:U2')
    dividends_cell = worksheet['T2']
    dividends_cell.value = 'Dividends (ttm)'
    dividends_cell.font = bold_font
    dividends_cell.fill = yellow_fill
    dividends_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')


    # Save the changes and close the Excel writer
    excel_writer.close()